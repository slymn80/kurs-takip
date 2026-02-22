import io
import os
import zipfile
import re
import json
from datetime import date, datetime
from flask import Blueprint, render_template, send_file, request, redirect, url_for, flash, current_app
from flask_login import login_required, current_user
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from sqlalchemy import or_
from ...models import Course, Enrollment, Student, Teacher, Organization, Certificate, AuditLog, CourseLedgerEntry, CourseExamResult, User
from ...extensions import db
from ...utils import serialize_json
from ...security import require_roles
from xml.sax.saxutils import escape as xml_escape


reports_bp = Blueprint("reports", __name__)


def _font_name():
    candidates = [
        ("Arial", r"C:\Windows\Fonts\arial.ttf"),
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    ]
    for name, path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
    return "Helvetica"


def _fmt_date(value):
    if value in (None, ""):
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value)


def _attache_name():
    attache = User.query.filter_by(role="attache", is_active=True).order_by(User.id.asc()).first()
    return attache.full_name if attache and attache.full_name else ""


def _build_report(report_type, filters=None):
    filters = filters or {}
    restrict_course_ids = None
    teacher_id_self = None
    if current_user.role == "teacher":
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        query = Course.query
        if teacher:
            teacher_id_self = teacher.id
            query = query.filter(Course.teacher_id == teacher.id)
        else:
            query = query.filter(Course.teacher_user_id == current_user.id)
        restrict_course_ids = [c.id for c in query.all()]

    if report_type == "courses":
        status = filters.get("course_status")
        query = Course.query
        if restrict_course_ids is not None:
            query = query.filter(Course.id.in_(restrict_course_ids))
        if status:
            query = query.filter(Course.status == status)
        rows = []
        for idx, c in enumerate(query.order_by(Course.created_at.desc()).all(), start=1):
            rows.append([idx, c.id, c.title, c.status, _fmt_date(c.start_date), _fmt_date(c.end_date), c.teacher.full_name if c.teacher else "-"])
        return "Kurs Listesi", ["Sıra No", "ID", "Kurs", "Durum", "Başlangıç", "Bitiş", "Öğretmen"], rows

    if report_type == "course_students":
        course_id = filters.get("course_id")
        query = Enrollment.query.join(Course).join(Student)
        if restrict_course_ids is not None:
            query = query.filter(Enrollment.course_id.in_(restrict_course_ids))
        if course_id:
            query = query.filter(Enrollment.course_id == course_id)
        rows = []
        for e in query.order_by(Course.title).all():
            rows.append([e.course.title, e.student.full_name, e.student.iin, e.student.phone or "-"])
        return "Kurs Bazında Öğrenci Listesi", ["Kurs", "Öğrenci", "IIN", "Telefon"], rows

    if report_type == "teachers":
        branch = filters.get("branch")
        query = Teacher.query
        if current_user.role == "teacher":
            query = query.filter(Teacher.user_id == current_user.id)
        if branch:
            query = query.filter(Teacher.branch == branch)
        rows = []
        for t in query.order_by(Teacher.full_name).all():
            rows.append([t.full_name, t.title, t.branch, t.phone, t.email])
        return "Öğretmen Listesi", ["Ad", "Unvan", "Branş", "Telefon", "E-posta"], rows

    if report_type == "organizations":
        org_id = filters.get("organization_id")
        query = Organization.query
        if restrict_course_ids is not None:
            query = query.join(Course, Course.organization_id == Organization.id).filter(Course.id.in_(restrict_course_ids))
        if org_id:
            query = query.filter(Organization.id == org_id)
        rows = []
        for o in query.order_by(Organization.name).all():
            rows.append([o.name, o.responsible_person, o.phone, o.email])
        return "Kurum Listesi", ["Kurum", "Sorumlu", "Telefon", "E-posta"], rows

    if report_type == "ended_courses":
        teacher_id = filters.get("teacher_id")
        query = Course.query.filter(Course.status == "ended")
        if restrict_course_ids is not None:
            query = query.filter(Course.id.in_(restrict_course_ids))
        if teacher_id:
            query = query.filter(Course.teacher_id == teacher_id)
        rows = []
        for c in query.order_by(Course.created_at.desc()).all():
            rows.append([c.id, c.title, _fmt_date(c.start_date), _fmt_date(c.end_date), c.teacher.full_name if c.teacher else "-"])
        return "Biten Kurslar", ["ID", "Kurs", "Başlangıç", "Bitiş", "Öğretmen"], rows

    if report_type == "dropped_courses":
        teacher_id = filters.get("teacher_id")
        query = Course.query.filter(Course.status == "dropped")
        if restrict_course_ids is not None:
            query = query.filter(Course.id.in_(restrict_course_ids))
        if teacher_id:
            query = query.filter(Course.teacher_id == teacher_id)
        rows = []
        for c in query.order_by(Course.created_at.desc()).all():
            rows.append([c.id, c.title, _fmt_date(c.start_date), _fmt_date(c.end_date), c.teacher.full_name if c.teacher else "-"])
        return "Yarım Kalan Kurslar", ["ID", "Kurs", "Başlangıç", "Bitiş", "Öğretmen"], rows

    if report_type == "teacher_courses":
        teacher_id = filters.get("teacher_id")
        rows = []
        if teacher_id:
            if current_user.role == "teacher" and teacher_id_self and teacher_id != teacher_id_self:
                return None, None, None
            for c in Course.query.filter(Course.teacher_id == teacher_id).order_by(Course.created_at.desc()).all():
                rows.append([c.title, c.status, _fmt_date(c.start_date), _fmt_date(c.end_date)])
            return "Öğretmene Atanan Kurslar", ["Kurs", "Durum", "Başlangıç", "Bitiş"], rows

        query = Course.query
        if restrict_course_ids is not None:
            query = query.filter(Course.id.in_(restrict_course_ids))
        for c in query.order_by(Course.created_at.desc()).all():
            rows.append([c.teacher.full_name if c.teacher else "-", c.title, c.status, _fmt_date(c.start_date), _fmt_date(c.end_date)])
        return "Tüm Öğretmenlerin Kursları", ["Öğretmen", "Kurs", "Durum", "Başlangıç", "Bitiş"], rows

    return None, None, None


def _pdf_report(title, headers, rows, meta_lines=None):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    font_name = _font_name()
    local_stamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    page_num = 1
    meta_lines = meta_lines or []

    def draw_footer(page_number):
        c.setFont(font_name, 8)
        c.drawString(40, 20, f"{local_stamp}")
        c.drawRightString(width - 40, 20, f"Sayfa {page_number}")

    def finish_page():
        nonlocal page_num
        draw_footer(page_num)
        c.showPage()
        page_num += 1

    def draw_signature_block():
        sig_y = 70
        c.setFont(font_name, 9)
        c.line(60, sig_y, 200, sig_y)
        c.line(width - 200, sig_y, width - 60, sig_y)
        c.drawString(60, sig_y - 12, "Öğretmen")
        right_center = width - 130
        right_name = _attache_name()
        if right_name:
            c.drawCentredString(right_center, sig_y + 10, right_name)
        c.drawCentredString(right_center, sig_y - 12, "Eğitim Ataşesi")

    c.setFont(font_name, 14)
    c.drawString(40, height - 40, title)
    c.setFont(font_name, 9)
    y = height - 64
    for line in meta_lines:
        c.drawString(40, y, line)
        y -= 12
    y -= 4
    col_width = max(80, int((width - 80) / max(1, len(headers))))

    def draw_row(values, y_pos, bold=False):
        if bold:
            c.setFont(font_name, 9)
        x = 40
        for val in values:
            text = str(val) if val is not None else ""
            c.drawString(x, y_pos, text[:30])
            x += col_width

    draw_row(headers, y, bold=True)
    y -= 18

    for row in rows:
        if y < 90:
            finish_page()
            c.setFont(font_name, 9)
            y = height - 60
            draw_row(headers, y, bold=True)
            y -= 18
        draw_row(row, y)
        y -= 14

    if y < 110:
        finish_page()
        c.setFont(font_name, 9)
        y = height - 60

    draw_signature_block()
    draw_footer(page_num)
    c.save()
    buffer.seek(0)
    return buffer


def _xlsx_report(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.8
    ws.oddFooter.left.text = "&D &T"
    ws.oddFooter.right.text = "Sayfa &P"

    ws.append(headers)
    for row in rows:
        ws.append(row)

    ws.append([])
    ws.append(["Öğretmen"])
    ws.append(["", "", "", _attache_name()])
    ws.append(["", "", "", "Eğitim Ataşesi"])

    from openpyxl.styles import Alignment
    name_row = ws.max_row - 0
    if ws.max_column >= 4:
        ws.merge_cells(start_row=name_row, start_column=4, end_row=name_row, end_column=ws.max_column)
        ws.cell(row=name_row, column=4).alignment = Alignment(horizontal="center")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _certificate_pdf(certificate):
    enrollment = certificate.enrollment
    course = enrollment.course
    student = enrollment.student
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    font_name = _font_name()

    # Border
    c.setStrokeColorRGB(0.78, 0.71, 0.56)
    c.setLineWidth(2)
    c.rect(30, 30, width - 60, height - 60)

    c.setFont(font_name, 20)
    c.drawCentredString(width / 2, height - 80, "SERTİFİKA")
    c.setFont(font_name, 12)
    c.drawCentredString(width / 2, height - 105, "Eğitim Ataşeliği")

    c.setFont(font_name, 14)
    c.drawCentredString(width / 2, height - 150, "Bu sertifika, aşağıdaki kursu başarıyla tamamlayan kursiyere verilmiştir.")

    c.setFont(font_name, 20)
    c.drawCentredString(width / 2, height - 210, student.full_name)

    c.setFont(font_name, 14)
    c.setFillColorRGB(0.89, 0.04, 0.09)
    c.drawCentredString(width / 2, height - 235, "BAŞARILI")
    c.setFillColorRGB(0.0, 0.0, 0.0)

    c.setFont(font_name, 12)
    course_name = course.title
    course_dates = f"{course.start_date.strftime('%d.%m.%Y')} - {course.end_date.strftime('%d.%m.%Y')}"
    c.drawCentredString(width / 2, height - 245, f"Kurs: {course_name}")
    c.drawCentredString(width / 2, height - 265, f"Tarih: {course_dates}")

    c.setFont(font_name, 10)
    c.drawString(50, 60, f"Sertifika No: {certificate.serial_no}")
    c.drawRightString(width - 50, 60, f"Düzenleme: {certificate.issued_at.strftime('%d.%m.%Y')}")

    # Signature lines
    c.line(80, 110, 260, 110)
    c.line(width - 260, 110, width - 80, 110)
    c.drawString(120, 95, "Eğitmen")
    right_name = _attache_name()
    if right_name:
        c.drawRightString(width - 120, 112, right_name)
    c.drawRightString(width - 120, 95, "Eğitim Ataşesi")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


def _attendance_certificate_payload():
    student_full_name = (request.args.get("student_full_name") or "").strip()
    hours_count = (request.args.get("hours_count") or "").strip()
    level = (request.args.get("level") or "").strip()
    cert_date = (request.args.get("cert_date") or "").strip()
    instructor_name = (request.args.get("instructor_name") or "").strip()
    if not cert_date:
        cert_date = datetime.now().strftime("%d.%m.%Y")
    return {
        "student_full_name": student_full_name,
        "hours_count": hours_count,
        "level": level,
        "cert_date": cert_date,
        "instructor_name": instructor_name
    }


def _attendance_certificate_template_path():
    project_root = os.path.abspath(os.path.join(current_app.root_path, ".."))
    candidates = [
        os.path.join(project_root, "Katılım Belgesi.docx"),
        os.path.join(project_root, "Katilim Belgesi.docx"),
        os.path.join(current_app.root_path, "static", "certificates", "katilim_belgesi.docx"),
        os.path.join(current_app.root_path, "static", "certificates", "Katılım Belgesi.docx"),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _replace_nearest_ellipsis(xml_text, anchor, value, search_forward=True, window=600):
    if not value:
        return xml_text
    idx = xml_text.find(anchor)
    if idx == -1:
        return xml_text
    if search_forward:
        segment_start = idx
        segment_end = min(len(xml_text), idx + window)
    else:
        segment_start = max(0, idx - window)
        segment_end = idx
    segment = xml_text[segment_start:segment_end]
    escaped_value = xml_escape(value)
    if value.startswith(" ") or value.endswith(" "):
        repl = f'<w:t xml:space="preserve">{escaped_value}</w:t>'
    else:
        repl = f"<w:t>{escaped_value}</w:t>"
    matches = list(re.finditer(r"<w:t[^>]*>(.*?)</w:t>", segment, flags=re.DOTALL))
    if not matches:
        return xml_text
    candidate_index = None
    for i, m in enumerate(matches):
        text_inner = (m.group(1) or "").strip()
        compact = text_inner.replace(" ", "")
        if "…" in compact or compact in {"..", "...", "....", "/../….", "/../../"}:
            candidate_index = i
            break
    if candidate_index is None:
        candidate_index = 0
    m = matches[candidate_index]
    segment2 = segment[:m.start()] + repl + segment[m.end():]
    return xml_text[:segment_start] + segment2 + xml_text[segment_end:]


def _replace_nearest_ellipsis_all(xml_text, anchor, value, search_forward=True, window=600):
    if not value:
        return xml_text
    out = xml_text
    start_pos = 0
    while True:
        idx = out.find(anchor, start_pos)
        if idx == -1:
            break
        if search_forward:
            segment_start = idx
            segment_end = min(len(out), idx + window)
            next_pos = idx + len(anchor)
        else:
            segment_start = max(0, idx - window)
            segment_end = idx
            next_pos = idx + len(anchor)
        segment = out[segment_start:segment_end]
        escaped_value = xml_escape(value)
        if value.startswith(" ") or value.endswith(" "):
            repl = f'<w:t xml:space="preserve">{escaped_value}</w:t>'
        else:
            repl = f"<w:t>{escaped_value}</w:t>"
        matches = list(re.finditer(r"<w:t[^>]*>(.*?)</w:t>", segment, flags=re.DOTALL))
        if not matches:
            start_pos = next_pos
            continue
        candidate_index = None
        for i, m in enumerate(matches):
            text_inner = (m.group(1) or "").strip()
            compact = text_inner.replace(" ", "")
            if "…" in compact or compact in {"..", "...", "....", "/../….", "/../../"}:
                candidate_index = i
                break
        if candidate_index is None:
            start_pos = next_pos
            continue
        m = matches[candidate_index]
        segment2 = segment[:m.start()] + repl + segment[m.end():]
        out = out[:segment_start] + segment2 + out[segment_end:]
        start_pos = max(0, segment_end - 1)
    return out


def _replace_docx_tokens_binary(template_path, replacements):
    """
    Replace token strings directly in DOCX XML parts.
    This preserves the font/style defined in the template because run properties stay intact.
    Also applies fallback replacements for legacy template marks (ellipsis-based placeholders).
    """
    out_stream = io.BytesIO()
    with zipfile.ZipFile(template_path, "r") as zin, zipfile.ZipFile(out_stream, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "word/document.xml":
                try:
                    xml_text = data.decode("utf-8")
                except UnicodeDecodeError:
                    zout.writestr(item, data)
                    continue
                updated = xml_text
                for key, value in replacements.items():
                    updated = updated.replace(key, xml_escape(value or ""))

                # Fallbacks for current template if explicit placeholders are absent.
                updated = _replace_nearest_ellipsis_all(updated, "Say", replacements.get("student_full_name", ""), search_forward=True)
                hours_text = (replacements.get("hours_count", "") or "").strip()
                level_text = (replacements.get("level", "") or "").strip()
                updated = _replace_nearest_ellipsis_all(
                    updated,
                    "<w:t>saatlik</w:t>",
                    f"{hours_text} " if hours_text else "",
                    search_forward=False
                )
                updated = _replace_nearest_ellipsis_all(
                    updated,
                    "zeyi</w:t>",
                    f" {level_text} " if level_text else "",
                    search_forward=False
                )
                if replacements.get("cert_date"):
                    updated = re.sub(
                        r"<w:t[^>]*>\s*/\s*\.\.\s*/\s*[^<]*</w:t>",
                        f"<w:t>{xml_escape(replacements['cert_date'])}</w:t>",
                        updated,
                        count=0,
                        flags=re.DOTALL
                    )
                updated = _replace_nearest_ellipsis_all(updated, "Kurs E", replacements.get("instructor_name", ""), search_forward=False)
                # "Kazanmıştır" sonrasında kalan gereksiz ".." parçasını temizle.
                updated = re.sub(r"<w:t[^>]*>\s*\.\.\s*</w:t>", '<w:t xml:space="preserve"> </w:t>', updated)

                data = updated.encode("utf-8")
            elif item.filename.endswith(".xml"):
                try:
                    xml_text = data.decode("utf-8")
                except UnicodeDecodeError:
                    zout.writestr(item, data)
                    continue
                updated = xml_text
                for key, value in replacements.items():
                    updated = updated.replace(key, xml_escape(value or ""))
                data = updated.encode("utf-8")
            zout.writestr(item, data)
    out_stream.seek(0)
    return out_stream


def _filled_attendance_docx_stream(data):
    template_path = _attendance_certificate_template_path()
    if not template_path:
        return None
    replacements = {
        "{{KURSIYER_ADI_SOYADI}}": data["student_full_name"],
        "{{SAAT_SAYISI}}": data["hours_count"],
        "{{DUZEY}}": data["level"],
        "{{TARIH}}": data["cert_date"],
        "{{KURS_EGITMENI}}": data["instructor_name"],
        "student_full_name": data["student_full_name"],
        "hours_count": data["hours_count"],
        "level": data["level"],
        "cert_date": data["cert_date"],
        "instructor_name": data["instructor_name"],
    }
    return _replace_docx_tokens_binary(template_path, replacements)


def _render_attendance_preview_html(data):
    stream = _filled_attendance_docx_stream(data)
    if not stream:
        return "<p>Şablon bulunamadı.</p>"
    qs = request.query_string.decode("utf-8", errors="ignore")
    src_url = url_for("reports.attendance_certificate_docx_source")
    if qs:
        src_url = f"{src_url}?{qs}"
    src_url_json = json.dumps(src_url)
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <style>
    html, body {{ margin: 0; padding: 0; background: #f7f7f7; }}
    #preview-root {{ padding: 20px; display: flex; justify-content: center; min-height: 95vh; }}
    #preview-root .docx-wrapper {{ background: #f7f7f7; }}
  </style>
  <script src="https://unpkg.com/docx-preview@0.3.3/dist/docx-preview.min.js"></script>
</head>
<body>
  <div id="preview-root">Belge yükleniyor...</div>
  <script>
    (async function() {{
      const root = document.getElementById("preview-root");
      try {{
        const response = await fetch({src_url_json}, {{ credentials: "same-origin" }});
        if (!response.ok) throw new Error("DOCX alınamadı");
        const buffer = await response.arrayBuffer();
        root.innerHTML = "";
        await window.docx.renderAsync(buffer, root, null, {{
          className: "docx",
          inWrapper: true,
          ignoreWidth: false,
          ignoreHeight: false,
          breakPages: true,
          ignoreFonts: false
        }});
      }} catch (_error) {{
        root.innerHTML = "<p>Önizleme oluşturulamadı.</p>";
      }}
    }})();
  </script>
</body>
</html>"""


def _academic_year_for_date(dt):
    if not dt:
        return None
    year = dt.year
    start_year = year if dt.month >= 8 else year - 1
    return f"{start_year}-{start_year + 1}"


def _attendance_percent_for_enrollment(enrollment):
    course = enrollment.course
    sessions = Session.query.filter_by(course_id=course.id, lesson_delivered=True).all()
    if not sessions:
        return 0.0
    session_ids = [s.id for s in sessions]
    total = len(session_ids)
    attended = Attendance.query.filter(
        Attendance.session_id.in_(session_ids),
        Attendance.student_id == enrollment.student_id,
        Attendance.status.in_(["present", "late", "excused"])
    ).count()
    return round((attended / total) * 100, 2) if total else 0.0


@reports_bp.route("/")
@login_required
@require_roles("teacher", "coordinator", "principal", "attache", "admin")
def index():
    teacher_id = request.args.get("teacher_id", type=int)
    course_status = request.args.get("course_status")
    course_id = request.args.get("course_id", type=int)
    branch = request.args.get("branch")
    organization_id = request.args.get("organization_id", type=int)
    ended_teacher_id = request.args.get("ended_teacher_id", type=int)
    dropped_teacher_id = request.args.get("dropped_teacher_id", type=int)

    courses_query = Course.query
    if current_user.role == "teacher":
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        if teacher:
            teacher_id = teacher.id
            courses_query = courses_query.filter(Course.teacher_id == teacher.id)
        else:
            courses_query = courses_query.filter(Course.teacher_user_id == current_user.id)
    if course_status:
        courses_query = courses_query.filter(Course.status == course_status)
    courses = courses_query.order_by(Course.created_at.desc()).all()

    course_students_query = Enrollment.query.join(Course).join(Student)
    if current_user.role == "teacher":
        course_students_query = course_students_query.filter(Course.id.in_([c.id for c in courses_query.all()]))
    if course_id:
        course_students_query = course_students_query.filter(Enrollment.course_id == course_id)
    course_students = course_students_query.order_by(Course.title).all()

    teachers_query = Teacher.query
    if current_user.role == "teacher":
        teachers_query = teachers_query.filter(Teacher.user_id == current_user.id)
    if branch:
        teachers_query = teachers_query.filter(Teacher.branch == branch)
    teachers = teachers_query.order_by(Teacher.full_name).all()

    organizations_query = Organization.query
    if current_user.role == "teacher":
        organizations_query = organizations_query.join(Course, Course.organization_id == Organization.id).filter(Course.id.in_([c.id for c in courses_query.all()]))
    if organization_id:
        organizations_query = organizations_query.filter(Organization.id == organization_id)
    organizations = organizations_query.order_by(Organization.name).all()

    ended_query = Course.query.filter(Course.status == "ended")
    if current_user.role == "teacher":
        ended_query = ended_query.filter(Course.id.in_([c.id for c in courses_query.all()]))
    if ended_teacher_id:
        ended_query = ended_query.filter(Course.teacher_id == ended_teacher_id)
    ended_courses = ended_query.order_by(Course.created_at.desc()).all()

    dropped_query = Course.query.filter(Course.status == "dropped")
    if current_user.role == "teacher":
        dropped_query = dropped_query.filter(Course.id.in_([c.id for c in courses_query.all()]))
    if dropped_teacher_id:
        dropped_query = dropped_query.filter(Course.teacher_id == dropped_teacher_id)
    dropped_courses = dropped_query.order_by(Course.created_at.desc()).all()

    teacher_courses = []
    if teacher_id:
        teacher_courses = Course.query.filter(Course.teacher_id == teacher_id).order_by(Course.created_at.desc()).all()
    else:
        if current_user.role == "teacher":
            teacher_courses = courses_query.order_by(Course.created_at.desc()).all()
        else:
            teacher_courses = Course.query.order_by(Course.created_at.desc()).all()

    all_courses = courses_query.order_by(Course.title).all()
    if current_user.role == "teacher":
        all_teachers = Teacher.query.filter_by(user_id=current_user.id).order_by(Teacher.full_name).all()
    else:
        all_teachers = Teacher.query.order_by(Teacher.full_name).all()

    return render_template(
        "reports/index.html",
        courses=courses,
        course_students=course_students,
        teachers=teachers,
        organizations=organizations,
        ended_courses=ended_courses,
        dropped_courses=dropped_courses,
        teacher_courses=teacher_courses,
        selected_teacher_id=teacher_id or "",
        selected_course_status=course_status or "",
        selected_course_id=course_id or "",
        selected_branch=branch or "",
        selected_organization_id=organization_id or "",
        selected_ended_teacher_id=ended_teacher_id or "",
        selected_dropped_teacher_id=dropped_teacher_id or "",
        all_courses=all_courses,
        all_teachers=all_teachers
    )


@reports_bp.route("/download/<report_type>")
@login_required
@require_roles("teacher", "coordinator", "principal", "attache", "admin")
def download(report_type):
    fmt = request.args.get("format", "pdf")
    filters = {
        "teacher_id": request.args.get("teacher_id", type=int),
        "course_status": request.args.get("course_status"),
        "course_id": request.args.get("course_id", type=int),
        "branch": request.args.get("branch"),
        "organization_id": request.args.get("organization_id", type=int)
    }
    if current_user.role == "teacher":
        teacher = Teacher.query.filter_by(user_id=current_user.id).first()
        allowed_query = Course.query
        if teacher:
            filters["teacher_id"] = teacher.id
            allowed_query = allowed_query.filter(Course.teacher_id == teacher.id)
        else:
            filters["teacher_id"] = None
            allowed_query = allowed_query.filter(Course.teacher_user_id == current_user.id)
        allowed_ids = {c.id for c in allowed_query.all()}
        if filters.get("course_id") and filters["course_id"] not in allowed_ids:
            filters["course_id"] = None
    title, headers, rows = _build_report(report_type, filters=filters)
    if not title:
        return "Not found", 404

    filename = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.{fmt}"

    if fmt == "xlsx":
        stream = _xlsx_report(headers, rows)
        return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    stream = _pdf_report(title, headers, rows)
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/pdf")


@reports_bp.route("/certificates")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def certificates():
    course_id = request.args.get("course_id", type=int)
    search = (request.args.get("q") or "").strip()
    courses = Course.query.filter(Course.status == "ended").order_by(Course.created_at.desc()).all()

    enrollments_query = (
        Enrollment.query
        .join(Course)
        .join(Student)
        .join(CourseExamResult, CourseExamResult.enrollment_id == Enrollment.id)
        .filter(
            Course.status == "ended",
            Enrollment.status == "active",
            CourseExamResult.passed.is_(True)
        )
    )
    if course_id:
        enrollments_query = enrollments_query.filter(Enrollment.course_id == course_id)
    if search:
        enrollments_query = enrollments_query.filter(
            (Student.full_name.ilike(f"%{search}%")) | (Student.iin.ilike(f"%{search}%"))
        )
    enrollments = enrollments_query.order_by(Student.full_name.asc()).all()

    certs = Certificate.query.filter(Certificate.enrollment_id.in_([e.id for e in enrollments])).all()
    cert_map = {c.enrollment_id: c for c in certs}
    results = CourseExamResult.query.filter(CourseExamResult.enrollment_id.in_([e.id for e in enrollments])).all()
    result_map = {r.enrollment_id: r for r in results}

    return render_template(
        "reports/certificates.html",
        courses=courses,
        selected_course_id=course_id or "",
        search=search,
        enrollments=enrollments,
        cert_map=cert_map,
        result_map=result_map
    )


@reports_bp.route("/certificates/<int:enrollment_id>/issue", methods=["POST"])
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def issue_certificate(enrollment_id):
    enrollment = Enrollment.query.get_or_404(enrollment_id)
    course = enrollment.course
    if course.status != "ended":
        return "Not allowed", 400
    existing = Certificate.query.filter_by(enrollment_id=enrollment.id).first()
    if existing:
        return redirect(url_for("reports.certificates", course_id=course.id))
    serial_no = f"ATAS-{datetime.utcnow().strftime('%Y%m')}-{enrollment.id:06d}"
    cert = Certificate(
        enrollment_id=enrollment.id,
        serial_no=serial_no,
        issued_by_user_id=current_user.id
    )
    db.session.add(cert)
    db.session.add(AuditLog(
        actor_user_id=current_user.id,
        actor_name_cached=current_user.full_name,
        actor_username_cached=current_user.username,
        action="create",
        entity_type="certificate",
        entity_id=0,
        after_json=serialize_json({"serial_no": serial_no})
    ))
    db.session.commit()
    return redirect(url_for("reports.certificate_pdf", certificate_id=cert.id))


@reports_bp.route("/certificates/<int:certificate_id>/pdf")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def certificate_pdf(certificate_id):
    cert = Certificate.query.get_or_404(certificate_id)
    stream = _certificate_pdf(cert)
    filename = f"certificate_{cert.serial_no}.pdf"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/pdf")


@reports_bp.route("/attendance-certificate")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def attendance_certificate():
    data = _attendance_certificate_payload()
    return render_template("reports/attendance_certificate.html", data=data)


@reports_bp.route("/attendance-certificate/preview")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def attendance_certificate_preview():
    data = _attendance_certificate_payload()
    html = _render_attendance_preview_html(data)
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


@reports_bp.route("/attendance-certificate/docx-source")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def attendance_certificate_docx_source():
    data = _attendance_certificate_payload()
    stream = _filled_attendance_docx_stream(data)
    if not stream:
        return "Template not found", 404
    return send_file(
        stream,
        as_attachment=False,
        download_name="katilim_belgesi_preview.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@reports_bp.route("/attendance-certificate/pdf")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def attendance_certificate_pdf():
    data = _attendance_certificate_payload()
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)
    font_name = _font_name()

    c.setStrokeColorRGB(0.78, 0.71, 0.56)
    c.setLineWidth(2)
    c.rect(30, 30, width - 60, height - 60)

    c.setFont(font_name, 22)
    c.drawCentredString(width / 2, height - 85, "KATILIM BELGESİ")
    c.setFont(font_name, 12)
    c.drawCentredString(width / 2, height - 110, "Almatı Eğitim Ataşeliği")

    c.setFont(font_name, 13)
    c.drawCentredString(width / 2, height - 155, "Aşağıda bilgileri yer alan kursiyerin kursa katılımı onaylanmıştır.")
    c.setFont(font_name, 20)
    c.drawCentredString(width / 2, height - 205, data["student_full_name"] or "........................................")

    c.setFont(font_name, 12)
    c.drawCentredString(
        width / 2,
        height - 245,
        f"Düzey: {data['level'] or '-'}   |   Saat: {data['hours_count'] or '-'}   |   Tarih: {data['cert_date'] or '-'}"
    )
    c.drawCentredString(width / 2, height - 270, f"Kurs Eğitmeni: {data['instructor_name'] or '-'}")

    c.line(90, 105, 290, 105)
    c.drawString(140, 90, "Kurs Eğitmeni")
    c.line(width - 290, 105, width - 90, 105)
    right_name = _attache_name()
    if right_name:
        c.drawRightString(width - 130, 107, right_name)
    c.drawRightString(width - 130, 90, "Eğitim Ataşesi")

    c.setFont(font_name, 9)
    c.drawString(40, 40, f"Oluşturma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    c.save()
    buffer.seek(0)
    filename = f"katilim_belgesi_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


@reports_bp.route("/attendance-certificate/docx")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def attendance_certificate_docx():
    template_path = _attendance_certificate_template_path()
    if not template_path:
        flash("Katılım Belgesi.docx şablonu bulunamadı.", "error")
        return redirect(url_for("reports.attendance_certificate"))

    data = _attendance_certificate_payload()
    stream = _filled_attendance_docx_stream(data)
    if not stream:
        flash("Katılım belgesi üretilemedi.", "error")
        return redirect(url_for("reports.attendance_certificate"))
    def _slug(v):
        txt = (v or "").strip().lower()
        tr_map = str.maketrans({
            "ç": "c", "ğ": "g", "ı": "i", "ö": "o", "ş": "s", "ü": "u",
            "Ç": "c", "Ğ": "g", "İ": "i", "Ö": "o", "Ş": "s", "Ü": "u"
        })
        txt = txt.translate(tr_map)
        txt = re.sub(r"[^a-z0-9]+", "_", txt)
        return txt.strip("_")

    student_part = _slug(data.get("student_full_name")) or "kursiyer"
    date_part = _slug(data.get("cert_date")) or datetime.now().strftime("%d_%m_%Y")
    hours_part = _slug(data.get("hours_count")) or "saat"
    level_part = _slug(data.get("level")) or "seviye"
    filename = f"{student_part}_{date_part}_{hours_part}_{level_part}.docx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@reports_bp.route("/attendance-certificate/template-docx")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def attendance_certificate_template_docx():
    template_path = _attendance_certificate_template_path()
    if not template_path:
        flash("Katılım Belgesi.docx şablonu bulunamadı.", "error")
        return redirect(url_for("reports.attendance_certificate"))
    return send_file(
        template_path,
        as_attachment=True,
        download_name="katilim_belgesi_sablon.docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@reports_bp.route("/course-ledger")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def course_ledger():
    course_id = request.args.get("course_id", type=int)
    academic_year = request.args.get("academic_year", "").strip()
    search = (request.args.get("q") or "").strip()

    courses = Course.query.filter(Course.status == "ended").order_by(Course.created_at.desc()).all()
    base_query = CourseLedgerEntry.query.outerjoin(Course, CourseLedgerEntry.course_id == Course.id).filter(
        or_(Course.id.is_(None), Course.status == "ended")
    )
    query = base_query
    if course_id:
        query = query.filter(CourseLedgerEntry.course_id == course_id)
    if academic_year:
        query = query.filter(CourseLedgerEntry.academic_year == academic_year)
    if search:
        query = query.filter(
            (CourseLedgerEntry.student_full_name_cached.ilike(f"%{search}%")) |
            (CourseLedgerEntry.student_iin_cached.ilike(f"%{search}%")) |
            (CourseLedgerEntry.course_title_cached.ilike(f"%{search}%"))
        )
    rows = query.order_by(CourseLedgerEntry.course_title_cached.asc(), CourseLedgerEntry.student_full_name_cached.asc()).all()
    years_from_entries = [y for (y,) in base_query.with_entities(CourseLedgerEntry.academic_year).distinct().all()]
    years_from_courses = {
        _academic_year_for_date(dt)
        for c in courses
        for dt in (c.start_date, c.end_date)
        if dt
    }
    academic_years = sorted({*years_from_entries, *years_from_courses} - {None, ""}, reverse=True)
    return render_template(
        "reports/course_ledger.html",
        courses=courses,
        academic_years=academic_years,
        selected_course_id=course_id or "",
        academic_year=academic_year,
        search=search,
        rows=rows
    )


@reports_bp.route("/course-ledger/generate", methods=["POST"])
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def generate_course_ledger():
    course_id = request.form.get("course_id", type=int)
    if not course_id:
        return redirect(url_for("reports.course_ledger"))
    course = Course.query.get_or_404(course_id)
    if course.status != "ended":
        flash("Kurs defteri yalnızca biten kurslar için oluşturulur.", "error")
        return redirect(url_for("reports.course_ledger", course_id=course.id))
    academic_year = _academic_year_for_date(course.start_date) or ""

    enrollments = Enrollment.query.filter_by(course_id=course.id).all()
    created = 0
    for enrollment in enrollments:
        existing = CourseLedgerEntry.query.filter_by(
            academic_year=academic_year,
            course_id=course.id,
            student_id=enrollment.student_id
        ).first()
        if existing:
            continue
        entry = CourseLedgerEntry(
            academic_year=academic_year,
            course_id=course.id,
            student_id=enrollment.student_id,
            teacher_user_id=course.teacher_user_id,
            course_title_cached=course.title,
            organization_name_cached=course.organization_name_cached or (course.organization.name if course.organization else None),
            location_name_cached=course.location_name_cached or (course.location.name if course.location else None),
            course_type_name_cached=course.course_type_name_cached or (course.course_type.name if course.course_type else None),
            teacher_name_cached=course.teacher_name_cached or (course.teacher.full_name if course.teacher else None),
            student_full_name_cached=enrollment.student.full_name,
            student_iin_cached=enrollment.student.iin,
            course_start_date=course.start_date,
            course_end_date=course.end_date,
            attendance_percent=_attendance_percent_for_enrollment(enrollment),
            result="Tamamladı"
        )
        db.session.add(entry)
        created += 1
    db.session.add(AuditLog(
        actor_user_id=current_user.id,
        actor_name_cached=current_user.full_name,
        actor_username_cached=current_user.username,
        action="generate",
        entity_type="course_ledger",
        entity_id=course.id,
        after_json=serialize_json({"created": created})
    ))
    db.session.commit()
    return redirect(url_for("reports.course_ledger", course_id=course.id, academic_year=academic_year))


@reports_bp.route("/course-ledger/<int:entry_id>/update", methods=["POST"])
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def update_course_ledger(entry_id):
    entry = CourseLedgerEntry.query.get_or_404(entry_id)
    entry.result = (request.form.get("result") or entry.result).strip()
    score = request.form.get("score")
    entry.score = int(score) if score and score.isdigit() else entry.score
    entry.notes = (request.form.get("notes") or "").strip()
    db.session.add(entry)
    db.session.add(AuditLog(
        actor_user_id=current_user.id,
        actor_name_cached=current_user.full_name,
        actor_username_cached=current_user.username,
        action="update",
        entity_type="course_ledger",
        entity_id=entry.id
    ))
    db.session.commit()
    return redirect(url_for("reports.course_ledger", course_id=entry.course_id, academic_year=entry.academic_year))


@reports_bp.route("/course-ledger/pdf")
@login_required
@require_roles("coordinator", "principal", "attache", "admin")
def course_ledger_pdf():
    course_id = request.args.get("course_id", type=int)
    academic_year = request.args.get("academic_year", "").strip()
    query = CourseLedgerEntry.query.outerjoin(Course, CourseLedgerEntry.course_id == Course.id).filter(
        or_(Course.id.is_(None), Course.status == "ended")
    )
    if course_id:
        query = query.filter(CourseLedgerEntry.course_id == course_id)
    if academic_year:
        query = query.filter(CourseLedgerEntry.academic_year == academic_year)
    rows = query.order_by(CourseLedgerEntry.course_title_cached.asc(), CourseLedgerEntry.student_full_name_cached.asc()).all()

    headers = ["Kursiyer", "IIN", "Kurs", "Öğretmen", "Dönem", "Devam%", "Sonuç", "Not"]
    body = []
    for r in rows:
        body.append([
            r.student_full_name_cached,
            r.student_iin_cached,
            r.course_title_cached,
            r.teacher_name_cached or "-",
            r.academic_year,
            r.attendance_percent,
            r.result,
            (r.notes or "")[:20]
        ])
    title = "Kurs Defteri"
    meta_lines = []
    if course_id:
        course = Course.query.get(course_id)
        meta_lines.append(f"Kurs: {course.title if course else 'Silinmiş Kurs'}")
    else:
        meta_lines.append("Kurs: Tümü")
    meta_lines.append(f"Eğitim Yılı: {academic_year or 'Tümü'}")
    meta_lines.append(f"Kayıt Sayısı: {len(rows)}")
    stream = _pdf_report(title, headers, body, meta_lines=meta_lines)
    filename = f"course_ledger_{academic_year or 'all'}.pdf"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/pdf")

